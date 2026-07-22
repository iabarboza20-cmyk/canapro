"""
Importa/actualiza asociados desde un Excel con la hoja "Asociados Activos"
(formato: DOCUMENTO, ASOCIADO, FECHA DE NACIMIENTO, EDAD, DIRECCION
RESIDENCIA, CORREO ELECTRONICO, TELEFONO CELULAR, MUNICIPIO DE TRABAJO,
INST EDUC, CARGO, CONYUGE, DOCUMENTO CONYUGE, BENEFICIARIO(S)..., ESTADO,
SEXO).

A diferencia de build_db.py, este script NO borra la base de datos:
  - Asociados que no existían: se insertan completos (con sus beneficiarios).
  - Asociados que ya existían: se actualizan solo los campos que vienen con
    dato en el Excel (si un campo viene vacío, se conserva el valor que ya
    había). Nunca se toca foto_url aquí, y los beneficiarios existentes NO
    se duplican: solo se agregan beneficiarios nuevos si el asociado no
    tenía ninguno registrado todavía.

Uso:
    python3 importar_asociados_activos.py ruta_al_excel.xlsx
"""
import sys
import sqlite3
import openpyxl

DB_PATH = "canaprosucre.db"

CAMPOS_ASOCIADO = [
    "nombre", "telefono", "email", "municipio", "estado",
    "municipio_trabajo", "institucion", "cargo", "edad", "sexo", "direccion",
]


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s != "0" else None


def importar(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Asociados Activos"]

    header_row = 2
    headers = [c.value for c in ws[header_row]]
    idx = {h: i for i, h in enumerate(headers) if h}

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    ben_pairs = [
        ("CONYUGE", "CONYUGE", "DOCUMENTO CONYUGE"),
        ("BENEFICIARIO 2", "BENEFICIARIO", "DOCUMENTO BENE"),
        ("BENEFICIARIO 3", "BENEFICIARIO 3", "DOCUMENTO BEN3"),
        ("BENEFICIARIO 4", "BENEFICIARIO 4", "DOCUMENTO BEN 4"),
        ("BENEFICIARIO 5", "BENEFICIARIO 5", "DOCUMENTO BE 5"),
    ]

    nuevos = 0
    actualizados = 0
    beneficiarios_agregados = 0

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        cedula = clean(row[idx["DOCUMENTO"]])
        nombre = clean(row[idx["ASOCIADO"]])
        if not cedula or not nombre:
            continue

        edad_raw = row[idx["EDAD"]]
        edad = str(int(edad_raw)) if isinstance(edad_raw, (int, float)) else clean(edad_raw)

        datos_nuevos = {
            "nombre": nombre,
            "telefono": clean(row[idx.get("TELEFONO CELULAR")]),
            "email": clean(row[idx.get("CORREO ELECTRONICO")]),
            "municipio": clean(row[idx.get("MUNICIPIO DE TRABAJO")]),
            "estado": clean(row[idx.get("ESTADO")]),
            "municipio_trabajo": clean(row[idx.get("MUNICIPIO DE TRABAJO")]),
            "institucion": clean(row[idx.get("INST EDUC")]),
            "cargo": clean(row[idx.get("CARGO")]),
            "edad": edad,
            "sexo": clean(row[idx.get("SEXO")]),
            "direccion": clean(row[idx.get("DIRECCION RESIDENCIA")]),
        }

        existente = conn.execute(
            "SELECT * FROM asociados WHERE cedula = ?", (cedula,)
        ).fetchone()

        if existente is None:
            conn.execute(
                "INSERT INTO asociados (cedula, " + ", ".join(CAMPOS_ASOCIADO) + ") "
                "VALUES (?, " + ", ".join(["?"] * len(CAMPOS_ASOCIADO)) + ")",
                [cedula] + [datos_nuevos[c] for c in CAMPOS_ASOCIADO],
            )
            nuevos += 1

            for parentesco, nombre_col, doc_col in ben_pairs:
                b_nombre = clean(row[idx.get(nombre_col)]) if nombre_col in idx else None
                b_doc = clean(row[idx.get(doc_col)]) if doc_col in idx else None
                if not b_nombre:
                    continue
                conn.execute(
                    "INSERT INTO beneficiarios (cedula_asociado, parentesco, nombre, documento) "
                    "VALUES (?, ?, ?, ?)",
                    (cedula, parentesco, b_nombre, b_doc),
                )
                beneficiarios_agregados += 1
        else:
            # Solo sobreescribe campos que vienen con dato nuevo; conserva
            # lo que ya había (incluye foto_url, que ni se toca).
            set_parts = []
            values = []
            for campo in CAMPOS_ASOCIADO:
                nuevo_valor = datos_nuevos[campo]
                if nuevo_valor is not None:
                    set_parts.append(f"{campo} = ?")
                    values.append(nuevo_valor)
            if set_parts:
                values.append(cedula)
                conn.execute(
                    f"UPDATE asociados SET {', '.join(set_parts)} WHERE cedula = ?",
                    values,
                )
                actualizados += 1

            # Si el asociado no tenía beneficiarios registrados, se cargan
            # los que trae este Excel. Si ya tenía, se deja igual (para no
            # duplicar).
            tiene_beneficiarios = conn.execute(
                "SELECT COUNT(*) c FROM beneficiarios WHERE cedula_asociado = ?", (cedula,)
            ).fetchone()["c"]
            if tiene_beneficiarios == 0:
                for parentesco, nombre_col, doc_col in ben_pairs:
                    b_nombre = clean(row[idx.get(nombre_col)]) if nombre_col in idx else None
                    b_doc = clean(row[idx.get(doc_col)]) if doc_col in idx else None
                    if not b_nombre:
                        continue
                    conn.execute(
                        "INSERT INTO beneficiarios (cedula_asociado, parentesco, nombre, documento) "
                        "VALUES (?, ?, ?, ?)",
                        (cedula, parentesco, b_nombre, b_doc),
                    )
                    beneficiarios_agregados += 1

    conn.commit()
    conn.close()

    print(f"Asociados nuevos creados: {nuevos}")
    print(f"Asociados existentes actualizados: {actualizados}")
    print(f"Beneficiarios agregados: {beneficiarios_agregados}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "ASOCIADOS_ACTIVOS_BASE_DE_DATOS.xlsx"
    importar(path)
