"""
Lee la hoja 'DATOS' del Excel de Canaprosucre y construye una base de datos
SQLite normalizada con dos tablas: asociados y beneficiarios.

Uso:
    python3 build_db.py ruta_al_excel.xlsx
"""
import sys
import re
import sqlite3
import openpyxl

DB_PATH = "canaprosucre.db"

SCHEMA = """
DROP TABLE IF EXISTS beneficiarios;
DROP TABLE IF EXISTS asociados;

CREATE TABLE asociados (
    cedula      TEXT PRIMARY KEY,
    nombre      TEXT NOT NULL,
    telefono    TEXT,
    email       TEXT,
    municipio   TEXT,
    estado      TEXT
);

CREATE TABLE beneficiarios (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    cedula_asociado TEXT NOT NULL REFERENCES asociados(cedula) ON DELETE CASCADE,
    parentesco  TEXT NOT NULL,   -- 'CONYUGE', 'BENEFICIARIO 2', 'BENEFICIARIO 3', ...
    nombre      TEXT NOT NULL,
    documento   TEXT
);

CREATE INDEX idx_beneficiarios_cedula ON beneficiarios(cedula_asociado);
"""


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def build(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["DATOS"]

    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA)
    cur = conn.cursor()

    # Pares (columna nombre beneficiario, columna documento) empezando por el conyuge
    ben_pairs = [
        ("CONYUGE", "CONYUGE", "DOCUMENTO"),
        ("BENEFICIARIO 2", "BEN 2", "DOCUMENTO B 2"),
        ("BENEFICIARIO 3", "BENEF 3", "DOCUME B3"),
        ("BENEFICIARIO 4", "BENEF 4", "DOCUME B4"),
        ("BENEFICIARIO 5", "BENEF 5", "DOCUME B5"),
        ("BENEFICIARIO 6", "BENEF 6", "DOCUME B6"),
        ("BENEFICIARIO 7", "BENEF 7", "DOCUME B7"),
    ]

    header_row = 3
    headers = [c.value for c in ws[header_row]]
    col_idx = {h: i for i, h in enumerate(headers) if h}

    n_asoc = 0
    n_ben = 0
    seen_cedulas = set()

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        cedula = clean(row[col_idx["CEDULA"]])
        nombre = clean(row[col_idx["ASOCIADO"]])
        if not cedula or not nombre:
            continue

        # Evita duplicados de cedula (se queda con el primero)
        if cedula in seen_cedulas:
            continue
        seen_cedulas.add(cedula)

        telefono = clean(row[col_idx["TELEFONO"]])
        email = clean(row[col_idx["CORREO EMAIL"]])
        municipio = clean(row[col_idx["MUNICIPIO"]])
        estado = clean(row[col_idx["ESTADO"]]) or "SIN DATO"

        cur.execute(
            "INSERT INTO asociados (cedula, nombre, telefono, email, municipio, estado) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (cedula, nombre, telefono, email, municipio, estado),
        )
        n_asoc += 1

        for parentesco, nombre_col, doc_col in ben_pairs:
            b_nombre = clean(row[col_idx[nombre_col]])
            b_doc = clean(row[col_idx[doc_col]])
            if not b_nombre:
                continue
            cur.execute(
                "INSERT INTO beneficiarios (cedula_asociado, parentesco, nombre, documento) "
                "VALUES (?, ?, ?, ?)",
                (cedula, parentesco, b_nombre, b_doc),
            )
            n_ben += 1

    conn.commit()
    conn.close()
    print(f"Listo. {n_asoc} asociados y {n_ben} beneficiarios cargados en {DB_PATH}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "BASE_DE_DATOS_ASOCIADOS_JULIO_15.xlsx"
    build(path)
